from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .models.cage import Cage
from .models.mouse import Mouse
from .models.user import User


HEADER_ALIASES = {
    "id#": "external_id",
    "id": "external_id",
    "new id# (retags)": "external_id",
    "new id# retags": "external_id",
    "retag": "external_id",
    "retags": "external_id",
    "mouse id": "external_id",
    "mouseid": "external_id",
    "mice id": "external_id",
    "miceid": "external_id",
    "sex": "gender",
    "gender": "gender",
    "dob": "dob",
    "date of birth": "dob",
    "birth date": "dob",
    "age (months)": "age_months",
    "age months": "age_months",
    "agemonths": "age_months",
    "genotype": "genotype",
    "geno": "genotype",
    "genotype reference #1": "genotype_reference_1",
    "genotype reference 1": "genotype_reference_1",
    "genotype ref #1": "genotype_reference_1",
    "genotype ref 1": "genotype_reference_1",
    "genotype reference #2": "genotype_reference_2",
    "genotype reference 2": "genotype_reference_2",
    "genotype ref #2": "genotype_reference_2",
    "genotype ref 2": "genotype_reference_2",
    "purpose": "purpose",
    "animal use (breeding/experimental)": "purpose",
    "animal use breeding/experimental": "purpose",
    "color": "color",
    "clr": "color",
    "barcodes": "cage_number",
    "barcode": "cage_number",
    "cage": "cage_number",
    "cage id": "cage_number",
    "cage number": "cage_number",
    "cage #": "cage_number",
    "male (m) (father)": "father",
    "male m father": "father",
    "father": "father",
    "female (f) (mother)": "mother",
    "female f mother": "mother",
    "mother": "mother",
    "toe id": "toe_id",
    "litter": "litter",
}

COMMON_FIELDS = {"external_id", "gender", "dob", "genotype", "cage_number"}


def normalize_header(value):
    text = str(value or "").strip().lower()
    for old, new in {
        "\n": " ",
        "\r": " ",
        "_": " ",
        "-": " ",
        ":": "",
    }.items():
        text = text.replace(old, new)
    return " ".join(text.split())


def map_header(value):
    normalized = normalize_header(value)
    compact = normalized.replace(" ", "")
    if normalized in HEADER_ALIASES:
        return HEADER_ALIASES[normalized]
    if compact in HEADER_ALIASES:
        return HEADER_ALIASES[compact]

    for alias, field in HEADER_ALIASES.items():
        if alias in normalized:
            return field
    return None


def normalize_cell(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text or None


def normalize_gender(value):
    text = normalize_cell(value)
    if not text:
        return "Unknown"
    lowered = text.lower()
    if lowered in {"m", "male"}:
        return "Male"
    if lowered in {"f", "female"}:
        return "Female"
    return text


def parse_dob(value):
    if value is None:
        return None
    if hasattr(value, "date"):
        return value.date()
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def calculate_age_months(dob):
    if dob is None:
        return None

    today = datetime.today().date()
    months = (today.year - dob.year) * 12 + today.month - dob.month
    if today.day < dob.day:
        months -= 1
    return str(max(months, 0))


def get_or_create_cage(db: Session, cage_number: str | None, user_id: int):
    if not cage_number:
        return None

    cage = (
        db.query(Cage)
        .filter(Cage.user_id == user_id, Cage.cage_number == cage_number)
        .first()
    )
    if cage:
        return cage

    cage = db.query(Cage).filter(Cage.cage_number == cage_number).first()
    if cage:
        return cage

    cage = Cage(cage_number=cage_number, user_id=user_id)
    db.add(cage)
    db.flush()
    return cage


def worksheet_rows_with_merged_values(sheet):
    merged_values = {}
    for merged_range in sheet.merged_cells.ranges:
        anchor_value = sheet.cell(merged_range.min_row, merged_range.min_col).value
        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            for column_index in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row_index, column_index)] = anchor_value

    rows = []
    for row_index in range(1, sheet.max_row + 1):
        row = []
        for column_index in range(1, sheet.max_column + 1):
            row.append(
                merged_values.get(
                    (row_index, column_index),
                    sheet.cell(row_index, column_index).value,
                )
            )
        rows.append(tuple(row))
    return rows


def find_header_row(rows):
    best_index = None
    best_headers = []
    best_score = 0

    for index, row in enumerate(rows[:25]):
        mapped_headers = [map_header(value) for value in row]
        score = len({header for header in mapped_headers if header})
        if score > best_score:
            best_index = index
            best_headers = mapped_headers
            best_score = score

    if best_index is None or best_score < 2:
        return None, []

    return best_index, best_headers


def looks_like_headerless_cage_list(rows):
    mouse_id_count = 0
    barcode_count = 0
    for row in rows:
        mouse_id = normalize_cell(row[1]) if len(row) > 1 else None
        barcode = normalize_cell(row[11]) if len(row) > 11 else None
        if mouse_id:
            mouse_id_count += 1
        if barcode:
            barcode_count += 1
    return mouse_id_count >= 5 and barcode_count >= 1


def headerless_cage_values(row, current_cage_number):
    cage_number = normalize_cell(row[11]) if len(row) > 11 else None
    if cage_number:
        current_cage_number = cage_number

    return (
        {
            "external_id": row[1] if len(row) > 1 else None,
            "gender": row[3] if len(row) > 3 else None,
            "color": row[4] if len(row) > 4 else None,
            "dob": row[5] if len(row) > 5 else None,
            "age_months": row[7] if len(row) > 7 else None,
            "genotype": row[8] if len(row) > 8 else None,
            "purpose": row[10] if len(row) > 10 else None,
            "cage_number": current_cage_number,
        },
        current_cage_number,
    )


def row_has_mouse_signal(values):
    return any(
        normalize_cell(values.get(field))
        for field in ("external_id", "gender", "genotype", "cage_number")
    ) or parse_dob(values.get("dob"))


def import_mice_from_xlsx(content: bytes, db: Session, current_user: User):
    workbook = load_workbook(BytesIO(content), data_only=True)
    imported = 0
    skipped = 0
    sheets_scanned = 0
    matched_fields = set()

    for sheet in workbook.worksheets:
        rows = worksheet_rows_with_merged_values(sheet)
        if not rows:
            continue
        sheets_scanned += 1

        header_index, headers = find_header_row(rows)
        headerless_cage_list = header_index is None and looks_like_headerless_cage_list(rows)
        current_cage_number = None

        if headerless_cage_list:
            data_rows = rows
            matched_fields.update(
                {"external_id", "gender", "color", "dob", "age_months", "genotype", "purpose", "cage_number"}
            )
        elif header_index is not None:
            data_rows = rows[header_index + 1 :]
            matched_fields.update(header for header in headers if header)
        else:
            continue

        for row in data_rows:
            if headerless_cage_list:
                values, current_cage_number = headerless_cage_values(row, current_cage_number)
            else:
                values = {}
                for index, header in enumerate(headers):
                    if not header or index >= len(row):
                        continue
                    current_value = row[index]
                    existing_value = values.get(header)
                    if normalize_cell(existing_value):
                        continue
                    values[header] = current_value
                cage_number = normalize_cell(values.get("cage_number"))
                if cage_number:
                    current_cage_number = cage_number
                elif current_cage_number:
                    values["cage_number"] = current_cage_number
            if not any(values.values()):
                continue
            has_external_id_column = "external_id" in matched_fields
            if has_external_id_column and not normalize_cell(values.get("external_id")):
                skipped += 1
                continue
            if not row_has_mouse_signal(values):
                skipped += 1
                continue

            dob = parse_dob(values.get("dob"))
            remark_parts = []
            if normalize_cell(values.get("color")):
                remark_parts.append(f"Color: {normalize_cell(values.get('color'))}")
            if normalize_cell(values.get("purpose")):
                remark_parts.append(f"Purpose: {normalize_cell(values.get('purpose'))}")
            if normalize_cell(values.get("father")):
                remark_parts.append(f"Father: {normalize_cell(values.get('father'))}")
            if normalize_cell(values.get("mother")):
                remark_parts.append(f"Mother: {normalize_cell(values.get('mother'))}")
            if normalize_cell(values.get("toe_id")):
                remark_parts.append(f"Toe ID: {normalize_cell(values.get('toe_id'))}")
            if normalize_cell(values.get("litter")):
                remark_parts.append(f"Litter: {normalize_cell(values.get('litter'))}")

            cage = get_or_create_cage(
                db,
                normalize_cell(values.get("cage_number")),
                current_user.id,
            )
            age_months = calculate_age_months(dob) or normalize_cell(values.get("age_months"))
            genotype = (
                normalize_cell(values.get("genotype"))
                or normalize_cell(values.get("genotype_reference_1"))
                or normalize_cell(values.get("genotype_reference_2"))
                or "Unknown"
            )
            mouse = Mouse(
                user_id=current_user.id,
                external_id=normalize_cell(values.get("external_id")),
                gender=normalize_gender(values.get("gender")),
                dob=dob,
                age_months=age_months,
                genotype=genotype,
                owner=current_user.username,
                remark="; ".join(remark_parts) or None,
            )
            if cage:
                mouse.cage = cage

            db.add(mouse)
            imported += 1

    db.commit()
    return {
        "imported": imported,
        "skipped": skipped,
        "sheets_scanned": sheets_scanned,
        "matched_fields": sorted(matched_fields),
        "missing_common_fields": sorted(COMMON_FIELDS - matched_fields),
    }
